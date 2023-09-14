from openpyxl import Workbook, load_workbook
import string


# Função para coletar e validar o login do usuário
def login(login_in):
    login = str(login_in).strip()
    if len(login) <= 15 and len(login) >= 5:
        return login
    else:
        return False


# Função para coletar e validar a senha do usuário
def senha(senha_in):
    especias = list(string.punctuation)
    senha = str(senha_in).strip()
    if len(senha) > 8:
        has_special = any(char in senha for char in especias)
        has_digit = any(char.isdigit() for char in senha)
        has_upper = any(char.isupper() for char in senha)

        if has_special and has_digit and has_upper:
            return senha
        else:
            return False
    else:
        return False


# Função para coletar e formatar o CPF do usuário
def cpf(cpf_in):
    cpf = str(cpf_in).strip()
    cpf_l = cpf.replace('.', '').replace('-', '').replace(',', '')
    if len(cpf_l) == 11:
        cpf1 = cpf_l[0:3]
        cpf2 = cpf_l[3:6]
        cpf3 = cpf_l[6:9]
        cpf4 = cpf_l[9:12]
        cpf_formatado = f'{cpf1}.{cpf2}.{cpf3}-{cpf4}'
        return cpf_formatado
    else:
        return False


def apresentacao(carta_in):
    carta = str(carta_in).strip()
    if 50 <= len(carta_in) <= 200:
        return carta
    else:
        return False


# Função principal para coletar dados e registrar no arquivo Excel
def cadastrar_dados(login_input, senha_input, cpf_input):
    try:
        wb = load_workbook('dados.xlsx')
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'LOGIN'
        ws['B1'] = 'SENHA'
        ws['C1'] = 'CPF'

    linha = ws.max_row + 1
    cadastrado = False

    # Verifica se o login ou CPF já existem no arquivo Excel
    for linha_atual in ws.iter_rows(min_row=2, values_only=True):
        if login_input == linha_atual[0] or cpf_input == linha_atual[2]:
            cadastrado = True
            return False

    # Se não estiver cadastrado, insere os dados no arquivo Excel
    if not cadastrado:
        ws.cell(row=linha, column=1, value=login_input)
        ws.cell(row=linha, column=2, value=senha_input)
        ws.cell(row=linha, column=3, value=cpf_input)
        wb.save('dados.xlsx')
        return True


def carta(carta_in):
    carta = str(carta_in).strip()
    if 50 <= len(carta_in
                 ) <= 200:
        return carta
    else:
        return False


def cadastrar_carta(login_input, carta_input):
    nome_arquivo = f"{login_input}.txt"
    try:
        with open(nome_arquivo, "w") as arquivo:
            arquivo.write(carta_input)
            print(f"Carta cadastrada com sucesso no arquivo {nome_arquivo}")
            return True
    except Exception as e:
        print(f"Erro ao cadastrar a carta: {str(e)}")
